import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import os
from pathlib import Path
import threading
from datetime import datetime
import logging
from pdf_processor import PDFProcessor


class LogWindow:
    def __init__(self, parent):
        self.window = tk.Toplevel(parent)
        self.window.title("Progreso del Procesamiento")
        self.window.geometry("600x650")
        self.window.resizable(True, True)
        self.window.configure(bg="#f8f9fa")

        # Center window
        self.window.transient(parent)
        self.window.grab_set()

        # Modern styling
        self.setup_styles()
        self.setup_ui()

    def setup_styles(self):
        self.style = ttk.Style()
        self.style.configure("Modern.TFrame", background="#f8f9fa")
        self.style.configure(
            "Header.TLabel",
            background="#f8f9fa",
            foreground="#2c3e50",
            font=("Segoe UI", 14, "bold"),
        )
        self.style.configure("Modern.TButton", padding=(20, 10), font=("Segoe UI", 10))

    def setup_ui(self):
        # Main container with padding
        main_frame = ttk.Frame(self.window, style="Modern.TFrame", padding=10)
        main_frame.grid(row=0, column=0, sticky="nsew")
        self.window.columnconfigure(0, weight=1)
        self.window.rowconfigure(0, weight=1)
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(1, weight=1)

        # Header with icon-like styling
        header_frame = ttk.Frame(main_frame, style="Modern.TFrame")
        header_frame.grid(row=0, column=0, sticky="ew", pady=(0, 20))

        ttk.Label(
            header_frame, text="📊 Progreso del Procesamiento", style="Header.TLabel"
        ).pack()

        # Separator line
        separator = ttk.Separator(main_frame, orient="horizontal")
        separator.grid(row=0, column=0, sticky="ew", pady=(40, 20))

        # Modern log area with better styling
        log_frame = ttk.Frame(main_frame, style="Modern.TFrame")
        log_frame.grid(row=1, column=0, sticky="nsew")
        log_frame.columnconfigure(0, weight=1)
        log_frame.rowconfigure(0, weight=1)

        # Text widget with modern appearance
        self.log_text = tk.Text(
            log_frame,
            wrap=tk.WORD,
            font=("Consolas", 10),
            bg="#ffffff",
            fg="#2c3e50",
            relief="flat",
            borderwidth=0,
            padx=15,
            pady=15,
        )

        # Modern scrollbar
        scrollbar = ttk.Scrollbar(
            log_frame, orient=tk.VERTICAL, command=self.log_text.yview
        )
        self.log_text.config(yscrollcommand=scrollbar.set)

        self.log_text.grid(row=0, column=0, sticky="nsew", padx=(0, 5))
        scrollbar.grid(row=0, column=1, sticky="ns")

        # Configure text styling
        self.log_text.tag_configure("INFO", foreground="#34495e")
        self.log_text.tag_configure(
            "WARNING", foreground="#f39c12", font=("Consolas", 10, "bold")
        )
        self.log_text.tag_configure(
            "ERROR", foreground="#e74c3c", font=("Consolas", 10, "bold")
        )
        self.log_text.tag_configure(
            "SUCCESS", foreground="#27ae60", font=("Consolas", 10, "bold")
        )
        self.log_text.tag_configure(
            "PROGRESS", foreground="#3498db", font=("Consolas", 10, "bold")
        )

        # Modern button area
        button_frame = ttk.Frame(main_frame, style="Modern.TFrame")
        button_frame.grid(row=2, column=0, pady=(20, 0))

        ttk.Button(
            button_frame,
            text="🗑️ Limpiar",
            style="Modern.TButton",
            command=self.clear_logs,
        ).grid(row=0, column=0, padx=(0, 15))

        ttk.Button(
            button_frame,
            text="✖️ Cerrar",
            style="Modern.TButton",
            command=self.close_window,
        ).grid(row=0, column=1)

        # Welcome message
        self.add_log("🚀 Iniciando procesamiento...", "PROGRESS")

    def add_log(self, message, level="INFO"):
        """Add a log message to the window with modern styling"""
        timestamp = datetime.now().strftime("%H:%M:%S")

        # Add emoji based on level
        emoji_map = {
            "INFO": "ℹ️",
            "WARNING": "⚠️",
            "ERROR": "❌",
            "SUCCESS": "✅",
            "PROGRESS": "🔄",
        }
        emoji = emoji_map.get(level, "ℹ️")

        formatted_message = f"[{timestamp}] {emoji} {message}\n"

        # Insert at end
        self.log_text.insert(tk.END, formatted_message, level)

        # Auto-scroll to bottom
        self.log_text.see(tk.END)

        # Update the window
        self.window.update_idletasks()

    def clear_logs(self):
        """Clear all logs"""
        self.log_text.delete(1.0, tk.END)
        self.add_log("Logs limpiados", "INFO")

    def close_window(self):
        """Close the log window"""
        self.window.grab_release()
        self.window.destroy()


class _GUILogHandler(logging.Handler):
    def __init__(self, gui, log_window=None):
        super().__init__()
        self.gui = gui
        self.log_window = log_window

    def emit(self, record):
        try:
            msg = self.format(record)
            # Send to main GUI
            self.gui.root.after(0, lambda: self.gui.log_message(msg))
            # Also send to log window if available
            if self.log_window:
                level_name = record.levelname
                if level_name == "ERROR":
                    tag = "ERROR"
                elif level_name == "WARNING":
                    tag = "WARNING"
                elif "completado" in msg.lower() or "éxito" in msg.lower():
                    tag = "SUCCESS"
                elif "procesando" in msg.lower() or "cargando" in msg.lower():
                    tag = "PROGRESS"
                else:
                    tag = "INFO"
                self.log_window.window.after(
                    0, lambda m=msg, t=tag: self.log_window.add_log(m, t)
                )
        except Exception:
            pass


class PDFProcessorGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("📄 Procesador de Soportes PDF")
        self.root.geometry("550x750")
        self.root.resizable(True, True)
        self.root.configure(bg="#f8f9fa")

        # Center window on screen
        self.root.after(100, self.center_window)

        # State variables
        self.pdf_file_path = tk.StringVar()
        self.output_dir_path = tk.StringVar(
            value=str(Path.cwd() / "soportes_separados")
        )
        self.extract_text_var = tk.BooleanVar(value=False)
        self.detailed_info_var = tk.BooleanVar(value=True)
        self.export_format_var = tk.StringVar(value="csv")
        self.initial_excel_path = tk.StringVar()
        self.search_col_var = tk.StringVar()
        self.rename_col_var = tk.StringVar()
        self.available_columns = []
        self.processing = False
        self.log_window = None

        self.setup_styles()
        self.setup_ui()
        self.setup_traces()
        self.update_process_button_state()

    def setup_styles(self):
        """Configure modern styling"""
        self.style = ttk.Style()

        # Try to use a modern theme
        try:
            self.style.theme_use("clam")
        except Exception:
            pass

        # Configure custom styles
        self.style.configure(
            "Card.TFrame", background="#ffffff", relief="flat", borderwidth=1
        )

        self.style.configure(
            "Header.TLabel",
            background="#f8f9fa",
            foreground="#2c3e50",
            font=("Segoe UI", 18, "bold"),
        )

        self.style.configure(
            "SubHeader.TLabel",
            background="#ffffff",
            foreground="#34495e",
            font=("Segoe UI", 11, "bold"),
        )

        self.style.configure(
            "Step.TLabel",
            background="#ffffff",
            foreground="#3498db",
            font=("Segoe UI", 12, "bold"),
        )

        self.style.configure(
            "Info.TLabel",
            background="#ffffff",
            foreground="#7f8c8d",
            font=("Segoe UI", 9),
        )

        self.style.configure(
            "Primary.TButton", font=("Segoe UI", 11, "bold"), padding=(20, 12)
        )

        self.style.configure(
            "Secondary.TButton", font=("Segoe UI", 10), padding=(15, 8)
        )

        self.style.configure("Modern.TEntry", padding=(10, 8), font=("Segoe UI", 10))

        # Additional styles for improved components
        self.style.configure(
            "Modern.TCheckbutton", font=("Segoe UI", 10), background="#ffffff"
        )

        self.style.configure(
            "Modern.Horizontal.TProgressbar",
            background="#3498db",
            troughcolor="#ecf0f1",
            borderwidth=0,
            lightcolor="#3498db",
            darkcolor="#3498db",
        )

    def setup_ui(self):
        """Setup the modern user interface with scrollable content"""
        # Configure grid weights
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)

        # Create a canvas and scrollbar for scrollable content
        canvas = tk.Canvas(self.root, bg="#f8f9fa", highlightthickness=0)
        scrollbar = ttk.Scrollbar(self.root, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)

        # Configure the canvas scrolling and responsive width
        def on_frame_configure(event):
            canvas.configure(scrollregion=canvas.bbox("all"))

        def on_canvas_configure(event):
            # Update the scroll region when canvas size changes
            canvas.configure(scrollregion=canvas.bbox("all"))
            # Make the frame fill the canvas width
            canvas_width = event.width
            canvas.itemconfig(window_id, width=canvas_width)

        scrollable_frame.bind("<Configure>", on_frame_configure)
        canvas.bind("<Configure>", on_canvas_configure)

        window_id = canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        # Place canvas and scrollbar
        canvas.grid(row=0, column=0, sticky="nsew")
        scrollbar.grid(row=0, column=1, sticky="ns")

        # Main container with responsive padding inside the scrollable frame
        main = ttk.Frame(scrollable_frame, padding=(15, 25, 15, 25))
        main.grid(row=0, column=0, sticky="ew")
        main.columnconfigure(0, weight=1)

        # Bind mousewheel to canvas for scroll functionality
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

        def _bind_to_mousewheel(event):
            canvas.bind_all("<MouseWheel>", _on_mousewheel)

        def _unbind_from_mousewheel(event):
            canvas.unbind_all("<MouseWheel>")

        canvas.bind("<Enter>", _bind_to_mousewheel)
        canvas.bind("<Leave>", _unbind_from_mousewheel)

        # Header
        header_frame = ttk.Frame(main)
        header_frame.grid(row=0, column=0, sticky="ew", pady=(0, 25))
        header_frame.columnconfigure(0, weight=1)

        ttk.Label(
            header_frame, text="📄 Procesador de Soportes PDF", style="Header.TLabel"
        ).grid(row=0, column=0)
        ttk.Label(
            header_frame,
            text="Herramienta para separar PDFs y renombrar archivos automáticamente",
            font=("Segoe UI", 11),
            foreground="#7f8c8d",
            background="#f8f9fa",
        ).grid(row=1, column=0, pady=(5, 0))

        # PDF selection
        self.create_pdf_section(main, 1)

        # Excel configuration
        self.create_excel_section(main, 2)

        # Output directory
        self.create_output_section(main, 3)

        # Additional options
        self.create_options_section(main, 4)

        # Action buttons
        self.create_action_section(main, 5)

        # Store references for cleanup
        self.canvas = canvas
        self.scrollable_frame = scrollable_frame

    def setup_traces(self):
        """Setup variable traces for UI updates"""
        self.rename_col_var.trace_add(
            "write",
            lambda *args: (self.update_preview(), self.update_process_button_state()),
        )
        self.search_col_var.trace_add(
            "write", lambda *args: self.update_process_button_state()
        )
        self.initial_excel_path.trace_add(
            "write", lambda *args: self.update_process_button_state()
        )
        self.pdf_file_path.trace_add(
            "write", lambda *args: self.update_process_button_state()
        )

    def center_window(self):
        """Center the window on the screen"""
        self.root.update_idletasks()
        width = self.root.winfo_width()
        height = self.root.winfo_height()
        x = (self.root.winfo_screenwidth() // 2) - (width // 2)
        y = (self.root.winfo_screenheight() // 2) - (height // 2)
        self.root.geometry(f"{width}x{height}+{x}+{y}")

    def create_pdf_section(self, parent, row):
        """Create PDF selection section"""
        pdf_card = ttk.Frame(parent, style="Card.TFrame", padding=20)
        pdf_card.grid(row=row, column=0, sticky="ew", pady=(0, 15))
        pdf_card.columnconfigure(1, weight=1)

        ttk.Label(pdf_card, text="🔶 PASO 1", style="Step.TLabel").grid(
            row=0, column=0, columnspan=3, sticky="w", pady=(0, 10)
        )
        ttk.Label(
            pdf_card, text="Seleccionar archivo PDF", style="SubHeader.TLabel"
        ).grid(row=1, column=0, columnspan=3, sticky="w", pady=(0, 15))

        ttk.Entry(
            pdf_card, textvariable=self.pdf_file_path, style="Modern.TEntry"
        ).grid(row=2, column=0, columnspan=2, sticky="ew", padx=(0, 15))
        ttk.Button(
            pdf_card,
            text="📁 Examinar",
            style="Secondary.TButton",
            command=self.select_pdf_file,
        ).grid(row=2, column=2)

    def create_excel_section(self, parent, row):
        """Create Excel configuration section"""
        excel_card = ttk.Frame(parent, style="Card.TFrame", padding=20)
        excel_card.grid(row=row, column=0, sticky="ew", pady=(0, 15))
        excel_card.columnconfigure(1, weight=1)

        ttk.Label(excel_card, text="🔶 PASO 2", style="Step.TLabel").grid(
            row=0, column=0, columnspan=3, sticky="w", pady=(0, 10)
        )
        ttk.Label(
            excel_card,
            text="Configurar Excel para búsqueda y renombrado",
            style="SubHeader.TLabel",
        ).grid(row=1, column=0, columnspan=3, sticky="w", pady=(0, 5))

        ttk.Label(
            excel_card,
            text="El Excel debe tener una columna con valores a buscar y otra con nombres para los archivos",
            style="Info.TLabel",
        ).grid(row=2, column=0, columnspan=3, sticky="w", pady=(0, 15))

        # Excel file selection
        ttk.Entry(
            excel_card, textvariable=self.initial_excel_path, style="Modern.TEntry"
        ).grid(row=3, column=0, columnspan=2, sticky="ew", padx=(0, 15))
        ttk.Button(
            excel_card,
            text="📊 Examinar",
            style="Secondary.TButton",
            command=self.select_initial_excel,
        ).grid(row=3, column=2)

        # Column selection
        col_frame = ttk.Frame(excel_card)
        col_frame.grid(row=4, column=0, columnspan=3, sticky="ew", pady=(20, 0))
        col_frame.columnconfigure(0, weight=1)
        col_frame.columnconfigure(1, weight=1)

        # Search column
        search_frame = ttk.Frame(col_frame)
        search_frame.grid(row=0, column=0, sticky="ew", padx=(0, 10))
        search_frame.columnconfigure(0, weight=1)
        ttk.Label(
            search_frame,
            text="🔍 Columna a buscar en PDF:",
            font=("Segoe UI", 10, "bold"),
        ).grid(row=0, column=0, sticky="w", pady=(0, 5))
        self.search_col_combo = ttk.Combobox(
            search_frame,
            textvariable=self.search_col_var,
            state="disabled",
            style="Modern.TEntry",
        )
        self.search_col_combo.grid(row=1, column=0, sticky="ew")

        # Rename column
        rename_frame = ttk.Frame(col_frame)
        rename_frame.grid(row=0, column=1, sticky="ew")
        rename_frame.columnconfigure(0, weight=1)
        ttk.Label(
            rename_frame,
            text="📝 Columna para nombrar archivo:",
            font=("Segoe UI", 10, "bold"),
        ).grid(row=0, column=0, sticky="w", pady=(0, 5))
        self.rename_col_combo = ttk.Combobox(
            rename_frame,
            textvariable=self.rename_col_var,
            state="disabled",
            style="Modern.TEntry",
        )
        self.rename_col_combo.grid(row=1, column=0, sticky="ew")

    def create_output_section(self, parent, row):
        """Create output directory section"""
        output_card = ttk.Frame(parent, style="Card.TFrame", padding=20)
        output_card.grid(row=row, column=0, sticky="ew", pady=(0, 15))
        output_card.columnconfigure(1, weight=1)

        ttk.Label(output_card, text="🔶 PASO 3", style="Step.TLabel").grid(
            row=0, column=0, columnspan=3, sticky="w", pady=(0, 10)
        )
        ttk.Label(
            output_card, text="Carpeta de destino", style="SubHeader.TLabel"
        ).grid(row=1, column=0, columnspan=3, sticky="w", pady=(0, 15))

        ttk.Entry(
            output_card, textvariable=self.output_dir_path, style="Modern.TEntry"
        ).grid(row=2, column=0, columnspan=2, sticky="ew", padx=(0, 15))
        ttk.Button(
            output_card,
            text="📂 Examinar",
            style="Secondary.TButton",
            command=self.select_output_dir,
        ).grid(row=2, column=2)

    def create_options_section(self, parent, row):
        """Create enhanced additional options section"""
        options_card = ttk.Frame(parent, style="Card.TFrame", padding=25)
        options_card.grid(row=row, column=0, sticky="ew", pady=(0, 15))
        options_card.columnconfigure(0, weight=1)

        ttk.Label(options_card, text="🔶 PASO 4", style="Step.TLabel").grid(
            row=0, column=0, sticky="w", pady=(0, 10)
        )
        ttk.Label(
            options_card,
            text="Configuración de procesamiento",
            style="SubHeader.TLabel",
        ).grid(row=1, column=0, sticky="w", pady=(0, 20))

        # Processing Options Section
        processing_frame = ttk.LabelFrame(
            options_card, text="⚙️ Opciones de procesamiento", padding=20
        )
        processing_frame.grid(row=2, column=0, sticky="ew", pady=(0, 15))
        processing_frame.columnconfigure(0, weight=1)

        # Extract text option with description
        extract_frame = ttk.Frame(processing_frame)
        extract_frame.grid(row=0, column=0, sticky="ew", pady=(0, 15))
        extract_frame.columnconfigure(0, weight=1)

        ttk.Checkbutton(
            extract_frame,
            text="📝 Extraer texto completo de páginas",
            variable=self.extract_text_var,
            style="Modern.TCheckbutton",
            command=self.on_extract_text_change,
        ).grid(row=0, column=0, sticky="w")

        extract_desc = ttk.Label(
            extract_frame,
            text="   💡 Extrae todo el contenido textual de cada página del PDF.\n   Útil para análisis posterior o búsquedas en el texto.",
            font=("Segoe UI", 9),
            foreground="#7f8c8d",
        )
        extract_desc.grid(row=1, column=0, sticky="w", pady=(5, 0))

        # Detailed info option with description
        detail_frame = ttk.Frame(processing_frame)
        detail_frame.grid(row=1, column=0, sticky="ew")
        detail_frame.columnconfigure(0, weight=1)

        ttk.Checkbutton(
            detail_frame,
            text="📊 Generar reporte detallado",
            variable=self.detailed_info_var,
            style="Modern.TCheckbutton",
            command=self.on_detailed_info_change,
        ).grid(row=0, column=0, sticky="w")

        detail_desc = ttk.Label(
            detail_frame,
            text="   💡 Crea un reporte con metadatos del PDF, estadísticas de procesamiento\n   y resumen de archivos generados.",
            font=("Segoe UI", 9),
            foreground="#7f8c8d",
        )
        detail_desc.grid(row=1, column=0, sticky="w", pady=(5, 0))

        # Export Format Section
        export_frame = ttk.LabelFrame(
            options_card, text="📤 Formato de exportación", padding=20
        )
        export_frame.grid(row=3, column=0, sticky="ew", pady=(0, 15))
        export_frame.columnconfigure(0, weight=1)
        export_frame.columnconfigure(1, weight=1)

        # Format selection
        format_left = ttk.Frame(export_frame)
        format_left.grid(row=0, column=0, sticky="ew", padx=(0, 15))
        format_left.columnconfigure(0, weight=1)

        ttk.Label(
            format_left, text="Seleccionar formato:", font=("Segoe UI", 10, "bold")
        ).grid(row=0, column=0, sticky="w", pady=(0, 8))
        format_combo = ttk.Combobox(
            format_left,
            textvariable=self.export_format_var,
            values=["csv", "xlsx"],
            state="readonly",
            style="Modern.TEntry",
            width=15,
        )
        format_combo.grid(row=1, column=0, sticky="w")
        format_combo.bind("<<ComboboxSelected>>", self.on_format_change)

        # Format descriptions
        format_right = ttk.Frame(export_frame)
        format_right.grid(row=0, column=1, sticky="ew")
        format_right.columnconfigure(0, weight=1)

        ttk.Label(
            format_right,
            text="Características del formato:",
            font=("Segoe UI", 10, "bold"),
        ).grid(row=0, column=0, sticky="w", pady=(0, 8))

        self.format_desc_var = tk.StringVar()
        self.update_format_description()

        format_desc_label = ttk.Label(
            format_right,
            textvariable=self.format_desc_var,
            font=("Segoe UI", 9),
            foreground="#34495e",
            justify="left",
        )
        format_desc_label.grid(row=1, column=0, sticky="w")

        # Preview section with enhanced styling
        preview_frame = ttk.LabelFrame(
            options_card, text="👁️ Vista previa del archivo", padding=20
        )
        preview_frame.grid(row=4, column=0, sticky="ew")
        preview_frame.columnconfigure(0, weight=1)

        ttk.Label(
            preview_frame, text="Patrón de nombres:", font=("Segoe UI", 10, "bold")
        ).grid(row=0, column=0, sticky="w", pady=(0, 8))

        self.preview_var = tk.StringVar(value="📄 soporte_<valor_extraído>.pdf")
        preview_entry = ttk.Entry(
            preview_frame,
            textvariable=self.preview_var,
            state="readonly",
            style="Modern.TEntry",
            font=("Segoe UI", 10, "bold"),
            foreground="#2c3e50",
        )
        preview_entry.grid(row=1, column=0, sticky="ew", pady=(0, 8))

        ttk.Label(
            preview_frame,
            text="💡 Los archivos se nombrarán automáticamente usando el valor\nencontrado en la columna seleccionada del PDF",
            font=("Segoe UI", 9),
            foreground="#7f8c8d",
        ).grid(row=2, column=0, sticky="w")

    def create_action_section(self, parent, row):
        """Create action buttons section"""
        action_card = ttk.Frame(parent, style="Card.TFrame", padding=25)
        action_card.grid(row=row, column=0, sticky="ew", pady=(0, 20))
        action_card.columnconfigure(0, weight=1)

        ttk.Label(action_card, text="🔶 PASO 5", style="Step.TLabel").grid(
            row=0, column=0, sticky="w", pady=(0, 10)
        )
        ttk.Label(
            action_card, text="Ejecutar procesamiento", style="SubHeader.TLabel"
        ).grid(row=1, column=0, sticky="w", pady=(0, 20))

        # Main action button
        main_button_frame = ttk.Frame(action_card)
        main_button_frame.grid(row=2, column=0, pady=(0, 20))

        self.process_button = ttk.Button(
            main_button_frame,
            text="🚀 PROCESAR PDF",
            style="Primary.TButton",
            command=self.start_processing,
        )
        self.process_button.pack(pady=(0, 10))

        # Secondary buttons in a more organized layout
        secondary_frame = ttk.Frame(action_card)
        secondary_frame.grid(row=3, column=0, sticky="ew", pady=(0, 20))
        secondary_frame.columnconfigure(0, weight=1)
        secondary_frame.columnconfigure(1, weight=1)
        secondary_frame.columnconfigure(2, weight=1)
        secondary_frame.columnconfigure(3, weight=1)

        # Row 1 of secondary buttons
        ttk.Button(
            secondary_frame,
            text="🗑️ Limpiar campos",
            style="Secondary.TButton",
            command=self.clear_fields,
        ).grid(row=0, column=0, padx=5, pady=2, sticky="ew")

        ttk.Button(
            secondary_frame,
            text="📁 Abrir carpeta",
            style="Secondary.TButton",
            command=lambda: self.ask_open_folder(Path(self.output_dir_path.get())),
        ).grid(row=0, column=1, padx=5, pady=2, sticky="ew")

        # Row 2 of secondary buttons
        ttk.Button(
            secondary_frame,
            text="📋 Ver progreso",
            style="Secondary.TButton",
            command=self.show_logs,
        ).grid(row=1, column=0, padx=5, pady=2, sticky="ew")

        ttk.Button(
            secondary_frame,
            text="❌ Salir",
            style="Secondary.TButton",
            command=self.root.quit,
        ).grid(row=1, column=1, padx=5, pady=2, sticky="ew")

        # Status section with better styling
        status_frame = ttk.LabelFrame(
            action_card, text="📊 Estado del procesamiento", padding=15
        )
        status_frame.grid(row=4, column=0, sticky="ew")
        status_frame.columnconfigure(0, weight=1)

        self.progress_var = tk.StringVar(
            value="Complete los pasos 1→2→3→4 antes de procesar"
        )
        status_label = ttk.Label(
            status_frame,
            textvariable=self.progress_var,
            font=("Segoe UI", 9),
            foreground="#34495e",
        )
        status_label.grid(row=0, column=0, sticky="w", pady=(0, 10))

        self.progress_bar = ttk.Progressbar(
            status_frame, mode="indeterminate", style="Modern.Horizontal.TProgressbar"
        )
        self.progress_bar.grid(row=1, column=0, sticky="ew")

    def clear_fields(self):
        """Clear all form fields"""
        self.pdf_file_path.set("")
        self.initial_excel_path.set("")
        self.search_col_var.set("")
        self.rename_col_var.set("")
        self.available_columns = []
        self.search_col_combo.configure(values=[])
        self.rename_col_combo.configure(values=[])
        self.search_col_combo.configure(state="disabled")
        self.rename_col_combo.configure(state="disabled")
        self.update_process_button_state()

    def select_initial_excel(self):
        file_path = filedialog.askopenfilename(
            title="Seleccionar archivo Excel",
            filetypes=[("Archivos Excel", "*.xlsx"), ("Todos los archivos", "*.*")],
            initialdir=str(Path.cwd()),
        )
        if not file_path:
            return
        self.initial_excel_path.set(file_path)
        self.log_message(f"Excel seleccionado: {Path(file_path).name}")
        try:
            from openpyxl import load_workbook

            wb = load_workbook(filename=file_path, read_only=True)
            ws = wb.active
            headers = [
                (str(c.value).strip() if c.value is not None else "") for c in ws[1]
            ]
            wb.close()
            self.available_columns = [h for h in headers if h]
            self.search_col_combo["values"] = self.available_columns
            self.rename_col_combo["values"] = self.available_columns
            self.search_col_combo.configure(state="readonly")
            self.rename_col_combo.configure(state="readonly")
            if self.available_columns:
                self.search_col_var.set(self.available_columns[0])
                self.rename_col_var.set(
                    self.available_columns[1]
                    if len(self.available_columns) > 1
                    else self.available_columns[0]
                )
            self.update_preview()
            self.log_message(
                f"Columnas detectadas: {', '.join(self.available_columns) if self.available_columns else 'Ninguna'}"
            )
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo leer el Excel: {e}")
            self.log_message(f"ERROR al leer Excel: {e}")

    def select_pdf_file(self):
        file_path = filedialog.askopenfilename(
            title="Seleccionar archivo PDF",
            filetypes=[("Archivos PDF", "*.pdf"), ("Todos los archivos", "*.*")],
            initialdir=str(Path.cwd()),
        )
        if file_path:
            self.pdf_file_path.set(file_path)
            self.log_message(f"Archivo seleccionado: {Path(file_path).name}")
            suggested_output = (
                Path(file_path).parent / f"soportes_{Path(file_path).stem}"
            )
            self.output_dir_path.set(str(suggested_output))

    def select_output_dir(self):
        dir_path = filedialog.askdirectory(
            title="Seleccionar carpeta de destino", initialdir=str(Path.cwd())
        )
        if dir_path:
            self.output_dir_path.set(dir_path)
            self.log_message(f"Carpeta de destino: {Path(dir_path).name}")

    def show_logs(self):
        """Show the log window"""
        if self.log_window is None or not self.log_window.window.winfo_exists():
            self.log_window = LogWindow(self.root)
        else:
            self.log_window.window.lift()
            self.log_window.window.focus()

    def log_message(self, message: str):
        ts = datetime.now().strftime("%H:%M:%S")
        print(f"[{ts}] {message}")
        self.root.update_idletasks()

    def validate_inputs(self) -> bool:
        if not self.pdf_file_path.get():
            messagebox.showerror("Error", "Por favor seleccione un archivo PDF")
            return False
        if not Path(self.pdf_file_path.get()).exists():
            messagebox.showerror("Error", "El archivo PDF seleccionado no existe")
            return False
        if Path(self.pdf_file_path.get()).suffix.lower() != ".pdf":
            messagebox.showerror("Error", "El archivo seleccionado no es un PDF válido")
            return False
        if not self.output_dir_path.get():
            messagebox.showerror("Error", "Por favor seleccione una carpeta de destino")
            return False
        if not self.initial_excel_path.get():
            messagebox.showerror(
                "Error - Paso 2 incompleto",
                "Debe seleccionar un archivo Excel que contenga los datos para búsqueda y renombrado.",
            )
            return False
        if not (self.search_col_var.get() and self.rename_col_var.get()):
            messagebox.showerror(
                "Error - Paso 3 incompleto",
                "Debe seleccionar AMBAS columnas:\n• Columna a buscar en el PDF\n• Columna para nombrar los archivos generados",
            )
            return False
        return True

    def update_process_button_state(self):
        can = bool(
            self.pdf_file_path.get()
            and self.output_dir_path.get()
            and self.initial_excel_path.get()
            and self.search_col_var.get()
            and self.rename_col_var.get()
        )
        try:
            self.process_button.configure(state=("normal" if can else "disabled"))
        except Exception:
            pass

    def on_extract_text_change(self):
        """Callback when extract text option changes"""
        if self.extract_text_var.get():
            print("Extracción de texto activada")
        else:
            print("Extracción de texto desactivada")

    def on_detailed_info_change(self):
        """Callback when detailed info option changes"""
        if self.detailed_info_var.get():
            print("Reporte detallado activado")
        else:
            print("Reporte detallado desactivado")

    def on_format_change(self, event=None):
        """Callback when export format changes"""
        self.update_format_description()

    def update_format_description(self):
        """Update the format description based on selected format"""
        format_val = self.export_format_var.get()
        if format_val == "csv":
            desc = "📋 CSV (Valores separados por comas)\n• Formato universal y ligero\n• Compatible con Excel y editores de texto\n• Ideal para análisis de datos"
        elif format_val == "xlsx":
            desc = "📊 XLSX (Excel nativo)\n• Formato nativo de Microsoft Excel\n• Soporte para fórmulas y formato\n• Mejor para presentaciones"
        else:
            desc = "Seleccione un formato para ver detalles"

        if hasattr(self, "format_desc_var"):
            self.format_desc_var.set(desc)

    def update_preview(self):
        val = (self.rename_col_var.get() or "valor").strip()
        self.preview_var.set(f"📄 soporte_{val}.pdf")

    def start_processing(self):
        if not self.validate_inputs():
            return
        if self.processing:
            messagebox.showwarning("Advertencia", "Ya hay un procesamiento en curso")
            return

        pdf_name = Path(self.pdf_file_path.get()).name
        out_dir = Path(self.output_dir_path.get()).name
        msg = [
            f"¿Procesar el archivo '{pdf_name}'?",
            f"Destino: {out_dir}",
            f"Extraer texto: {'Sí' if self.extract_text_var.get() else 'No'}",
            f"Info detallada: {'Sí' if self.detailed_info_var.get() else 'No'}",
            f"Exportación: {self.export_format_var.get().upper()}",
            f"Excel: {Path(self.initial_excel_path.get()).name}",
            f"Buscar en: {self.search_col_var.get()} | Renombrar con: {self.rename_col_var.get()}",
        ]
        if not messagebox.askyesno("Confirmar procesamiento", "\n".join(msg)):
            return

        # Open the log window before starting processing
        self.log_window = LogWindow(self.root)

        self.processing = True
        self.process_button.config(state="disabled")
        self.progress_bar.start()
        threading.Thread(target=self.process_pdf, daemon=True).start()

    def process_pdf(self):
        try:
            self.progress_var.set("Iniciando procesamiento...")
            self.log_message("=" * 50)
            self.log_message("INICIANDO PROCESAMIENTO")
            self.log_message("=" * 50)

            # Add detailed info to log window
            if self.log_window:
                self.log_window.add_log(
                    f"📄 Archivo PDF: {Path(self.pdf_file_path.get()).name}", "PROGRESS"
                )
                self.log_window.add_log(
                    f"📁 Carpeta destino: {self.output_dir_path.get()}", "PROGRESS"
                )
                self.log_window.add_log(
                    f"📊 Excel: {Path(self.initial_excel_path.get()).name}", "PROGRESS"
                )
                self.log_window.add_log(
                    f"🔍 Columna búsqueda: {self.search_col_var.get()}", "PROGRESS"
                )
                self.log_window.add_log(
                    f"📝 Columna renombrado: {self.rename_col_var.get()}", "PROGRESS"
                )
                self.log_window.add_log(
                    f"📤 Formato exportación: {self.export_format_var.get().upper()}",
                    "PROGRESS",
                )

            self.progress_var.set("Cargando Excel y preparando mapeo...")
            self.log_message("Cargando Excel y preparando mapeo...")

            mapping = (self.search_col_var.get(), self.rename_col_var.get())

            logger = logging.getLogger("pdf_processor")
            logger.setLevel(logging.INFO)
            if not hasattr(self, "_gui_handler"):
                self._gui_handler = _GUILogHandler(self, self.log_window)
                self._gui_handler.setLevel(logging.INFO)
                self._gui_handler.setFormatter(logging.Formatter("%(message)s"))
                logger.addHandler(self._gui_handler)

            processor = PDFProcessor(
                self.pdf_file_path.get(),
                self.output_dir_path.get(),
                export_format=self.export_format_var.get(),
                initial_excel_path=self.initial_excel_path.get(),
                mapping_columns=mapping,
            )

            self.progress_var.set("Validando archivo PDF...")
            if self.log_window:
                self.log_window.add_log("🔍 Validando archivo PDF...", "PROGRESS")
            processor.validate_input()

            self.progress_var.set("Procesando páginas...")
            if self.log_window:
                self.log_window.add_log(
                    "📄 Iniciando separación de páginas...", "PROGRESS"
                )
            created_files = processor.separate_pages()

            if self.extract_text_var.get():
                self.progress_var.set("Extrayendo texto...")
                self.log_message("Extrayendo texto de las páginas...")
                if self.log_window:
                    self.log_window.add_log(
                        "📝 Extrayendo texto de páginas...", "PROGRESS"
                    )
                processor.extract_text_from_pages()

            if self.detailed_info_var.get():
                self.progress_var.set("Generando información detallada...")
                self.log_message("Extrayendo información detallada de soportes...")
                if self.log_window:
                    self.log_window.add_log(
                        "📊 Generando información detallada...", "PROGRESS"
                    )
                import fitz

                pdf_doc = fitz.open(self.pdf_file_path.get())
                metadata = processor.extract_metadata(pdf_doc)
                pdf_doc.close()
                processor.create_detailed_summary_report(metadata, created_files)

            self.progress_var.set("¡Procesamiento completado exitosamente!")
            self.log_message("=" * 50)
            self.log_message("PROCESAMIENTO COMPLETADO")
            self.log_message("=" * 50)
            self.log_message(f"Archivos creados: {len(created_files)}")
            self.log_message(f"Ubicación: {processor.output_dir}")

            if self.log_window:
                self.log_window.add_log("=" * 50, "SUCCESS")
                self.log_window.add_log(
                    "✅ PROCESAMIENTO COMPLETADO EXITOSAMENTE", "SUCCESS"
                )
                self.log_window.add_log("=" * 50, "SUCCESS")
                self.log_window.add_log(
                    f"📁 Total de archivos creados: {len(created_files)}", "SUCCESS"
                )
                self.log_window.add_log(
                    f"📂 Ubicación: {processor.output_dir}", "SUCCESS"
                )

            def _after_success():
                if messagebox.askyesno(
                    "Éxito",
                    f"Procesamiento completado!\n\nArchivos creados: {len(created_files)}\nUbicación: {processor.output_dir}\n\n¿Desea abrir la carpeta de destino?",
                ):
                    self.ask_open_folder(processor.output_dir)

            self.root.after(0, _after_success)
        except Exception as e:
            self.progress_var.set("Error en el procesamiento")
            self.log_message(f"ERROR: {e}")
            if self.log_window:
                self.log_window.add_log(f"❌ ERROR: {e}", "ERROR")
            err_msg = str(e)
            self.root.after(0, lambda m=err_msg: messagebox.showerror("Error", m))
        finally:
            self.processing = False
            self.root.after(0, lambda: self.process_button.config(state="normal"))
            self.root.after(0, lambda: self.progress_bar.stop())
            try:
                if hasattr(self, "_gui_handler"):
                    logging.getLogger("pdf_processor").removeHandler(self._gui_handler)
                    del self._gui_handler
            except Exception:
                pass

    def ask_open_folder(self, folder_path):
        import sys
        import subprocess

        folder = Path(folder_path) if not isinstance(folder_path, Path) else folder_path
        if not folder.exists():
            if not messagebox.askyesno(
                "Carpeta no existe", "La carpeta de destino no existe. ¿Desea crearla?"
            ):
                return
            try:
                folder.mkdir(parents=True, exist_ok=True)
            except Exception as e:
                messagebox.showerror("Error", f"No se pudo crear la carpeta: {e}")
                return
        try:
            if sys.platform.startswith("win"):
                os.startfile(str(folder))
            elif sys.platform == "darwin":
                subprocess.call(["open", str(folder)])
            else:
                subprocess.call(["xdg-open", str(folder)])
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo abrir la carpeta: {e}")


def main():
    root = tk.Tk()
    app = PDFProcessorGUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()
