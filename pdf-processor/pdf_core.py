from pathlib import Path
import fitz
import re
from typing import Dict, Optional, Tuple, List
from logger_config import get_logger

try:
    from openpyxl import load_workbook
    from openpyxl import Workbook
except Exception:
    load_workbook = None
    Workbook = None


class PDFProcessor:
    def __init__(
        self,
        input_pdf_path,
        output_dir=None,
        export_format: str = "csv",
        initial_excel_path: Optional[str] = None,
        mapping_columns: Optional[Tuple[str, str]] = None,
    ):
        self.input_pdf_path = Path(input_pdf_path)
        self.output_dir = (
            Path(output_dir)
            if output_dir
            else self.input_pdf_path.parent / "soportes_separados"
        )
        self.export_format = (export_format or "csv").lower()
        if self.export_format not in ("csv", "xlsx"):
            self.export_format = "csv"
        self.initial_excel_path = (
            Path(initial_excel_path) if initial_excel_path else None
        )
        self.mapping_columns = mapping_columns
        self.search_to_rename_map: Dict[str, str] = {}
        self.logger = get_logger("pdf_processor")

        if self.initial_excel_path and self.mapping_columns:
            try:
                self.search_to_rename_map = self.load_excel_mapping(
                    self.initial_excel_path,
                    self.mapping_columns[0],
                    self.mapping_columns[1],
                )
                self.logger.info(
                    f"Mapa de {len(self.search_to_rename_map)} valores cargado desde Excel"
                )
                self.logger.info(f"Columna de búsqueda: '{self.mapping_columns[0]}'")
                self.logger.info(f"Columna de renombrado: '{self.mapping_columns[1]}'")
                self._log_mapping_summary()
            except Exception as e:
                self.logger.error(f"No se pudo cargar el Excel inicial: {e}")
                self.search_to_rename_map = {}

    def _log_mapping_summary(self):
        """Muestra un resumen estadístico del mapeo cargado, similar a pandas.info()"""
        if not self.search_to_rename_map:
            return

        total_entries = len(self.search_to_rename_map)

        search_lengths = [len(str(key)) for key in self.search_to_rename_map.keys()]
        min_search_len = min(search_lengths) if search_lengths else 0
        max_search_len = max(search_lengths) if search_lengths else 0
        avg_search_len = (
            sum(search_lengths) / len(search_lengths) if search_lengths else 0
        )

        rename_lengths = [len(str(val)) for val in self.search_to_rename_map.values()]
        min_rename_len = min(rename_lengths) if rename_lengths else 0
        max_rename_len = max(rename_lengths) if rename_lengths else 0
        avg_rename_len = (
            sum(rename_lengths) / len(rename_lengths) if rename_lengths else 0
        )

        self.logger.info("=== RESUMEN DEL MAPEO EXCEL ===")
        self.logger.info(f"Total de entradas: {total_entries}")
        self.logger.info(
            f"Longitud claves búsqueda: min={min_search_len}, max={max_search_len}, promedio={avg_search_len:.1f}"
        )
        self.logger.info(
            f"Longitud valores renombrado: min={min_rename_len}, max={max_rename_len}, promedio={avg_rename_len:.1f}"
        )

        sample_items = list(self.search_to_rename_map.items())[:5]
        self.logger.info("=== EJEMPLOS DE MAPEO ===")
        for i, (search_val, rename_val) in enumerate(sample_items, 1):
            self.logger.info(f"  {i}. '{search_val}' -> '{rename_val}'")

        if total_entries > 5:
            self.logger.info(f"  ... y {total_entries - 5} entradas más")
        self.logger.info("=" * 30)

    def load_excel_mapping(
        self, excel_path: Path, search_col_name: str, rename_col_name: str
    ) -> Dict[str, str]:
        if load_workbook is None:
            raise RuntimeError(
                "openpyxl no está instalado. No se puede leer el Excel inicial."
            )
        if not excel_path.exists():
            raise FileNotFoundError(f"Excel no encontrado: {excel_path}")

        wb = load_workbook(filename=str(excel_path), read_only=True, data_only=True)
        ws = wb.active

        headers: List[str] = []
        for cell in ws[1]:
            cell_value = cell.value
            if cell_value is not None:
                headers.append(str(cell_value).strip())
            else:
                headers.append("")

        self.logger.debug(f"Headers encontrados: {headers}")

        norm_headers = {
            h.strip().casefold(): i for i, h in enumerate(headers) if h.strip()
        }
        s_key = (search_col_name or "").strip().casefold()
        r_key = (rename_col_name or "").strip().casefold()

        if s_key not in norm_headers or r_key not in norm_headers:
            wb.close()
            available_headers = [h for h in headers if h.strip()]
            raise ValueError(
                "No se encontraron las columnas seleccionadas en el Excel. "
                f"Disponibles: {', '.join([repr(h) for h in available_headers])}"
            )

        search_idx = norm_headers[s_key]
        rename_idx = norm_headers[r_key]

        mapping: Dict[str, str] = {}
        processed_rows = 0
        empty_rows = 0

        for row_num, row in enumerate(ws.iter_rows(min_row=2), start=2):
            processed_rows += 1

            s_val = row[search_idx].value if len(row) > search_idx else None
            r_val = row[rename_idx].value if len(row) > rename_idx else None

            if s_val is None and r_val is None:
                empty_rows += 1
                continue

            if s_val is None or r_val is None:
                self.logger.warning(
                    f"Fila {row_num}: Valor faltante - Búsqueda: '{s_val}', Renombrado: '{r_val}'"
                )
                continue

            s = str(s_val).strip()
            r = str(r_val).strip()

            if not s or not r:
                self.logger.warning(
                    f"Fila {row_num}: Valores vacíos después de normalizar - Búsqueda: '{s}', Renombrado: '{r}'"
                )
                continue

            if s in mapping:
                self.logger.warning(
                    f"Fila {row_num}: Valor de búsqueda duplicado '{s}' - manteniendo primer valor '{mapping[s]}'"
                )
                continue

            mapping[s] = r
            self.logger.debug(f"Fila {row_num}: Mapeado '{s}' -> '{r}'")

        wb.close()

        self.logger.info(
            f"Procesadas {processed_rows} filas, {empty_rows} vacías, {len(mapping)} mapeos válidos"
        )

        if not mapping:
            raise ValueError(
                "No se pudo crear ningún mapeo válido desde el Excel. "
                "Verifique que las columnas tengan datos."
            )

        return mapping

    def _sanitize_filename(self, name: str) -> str:
        """Sanitiza un nombre para que sea válido como nombre de archivo"""
        if not name:
            return "unnamed"
        sanitized = re.sub(r'[<>:"/\\|?*]', "_", str(name))
        sanitized = re.sub(r"\s+", "_", sanitized)
        sanitized = sanitized.strip("._")
        return sanitized[:100] if sanitized else "unnamed"

    def _unique_path(self, base_dir: Path, filename: str) -> Path:
        """Genera una ruta única en el directorio base"""
        path = base_dir / filename
        if not path.exists():
            return path

        name_part = path.stem
        extension = path.suffix
        counter = 1

        while True:
            new_name = f"{name_part}_{counter}{extension}"
            new_path = base_dir / new_name
            if not new_path.exists():
                return new_path
            counter += 1

    def _normalize_text_for_search(self, text: str) -> str:
        if not text:
            return ""
        normalized = " ".join(text.split())
        normalized = normalized.lower()
        normalized = re.sub(r"[^\w\s]", " ", normalized)
        normalized = " ".join(normalized.split())
        return normalized

    def _extract_value_after_header(self, text: str, header_name: str) -> Optional[str]:
        """Extrae el valor que viene después del nombre del header en el texto"""
        if not text or not header_name:
            return None

        pattern = re.escape(header_name)
        match = re.search(pattern, text, re.IGNORECASE)

        if not match:
            header_words = header_name.split()
            if len(header_words) > 1:
                pattern = r"\s+".join([re.escape(word) for word in header_words])
                match = re.search(pattern, text, re.IGNORECASE)

        if not match:
            return None

        text_after_header = text[match.end() :]
        lines = text_after_header.split("\n")

        for line in lines[:3]:
            clean_line = line.strip()
            if not clean_line:
                continue

            clean_line = re.sub(r"^[:\-\s]+", "", clean_line)

            if clean_line:
                value_match = re.search(r"([A-Za-z0-9][A-Za-z0-9\-_]*)", clean_line)
                if value_match:
                    extracted_value = value_match.group(1)
                    self.logger.debug(
                        f"Valor extraído después de '{header_name}': '{extracted_value}' de la línea: '{clean_line.strip()}'"
                    )
                    return extracted_value

        return None

    def _find_rename_in_text(self, text: str) -> Optional[str]:
        if not self.search_to_rename_map:
            self.logger.debug("No hay mapeo cargado desde Excel")
            return None

        self.logger.debug("=== INICIANDO BÚSQUEDA DE COINCIDENCIAS ===")
        self.logger.debug(f"Texto PDF original (primeros 500 chars): {text[:500]}...")

        normalized_pdf_text = self._normalize_text_for_search(text)
        self.logger.debug(
            f"Texto PDF normalizado (primeros 300 chars): {normalized_pdf_text[:300]}..."
        )

        extracted_values = []

        self.logger.debug("=== MÉTODO 1: Búsqueda directa de valores del Excel ===")
        for i, key in enumerate(
            sorted(
                self.search_to_rename_map.keys(),
                key=lambda k: len(str(k)),
                reverse=True,
            )[:10]
        ):
            if not key:
                continue

            normalized_key = self._normalize_text_for_search(str(key))
            self.logger.debug(f"  Buscando: '{key}' (normalizado: '{normalized_key}')")

            if not normalized_key:
                self.logger.debug("    -> Clave vacía después de normalizar")
                continue

            if normalized_key in normalized_pdf_text:
                self.logger.info(
                    f"✅ COINCIDENCIA DIRECTA ENCONTRADA: '{key}' -> '{self.search_to_rename_map[key]}'"
                )
                return self.search_to_rename_map[key]
            else:
                self.logger.debug("    -> No encontrado en texto PDF")

        self.logger.debug("=== MÉTODO 2: Búsqueda por header de columna ===")
        if self.mapping_columns and len(self.mapping_columns) >= 2:
            for col_name in self.mapping_columns:
                if col_name and col_name.strip():
                    self.logger.debug(
                        f"Buscando header de columna '{col_name}' en el texto del PDF"
                    )

                    extracted_value = self._extract_value_after_header(text, col_name)

                    if extracted_value:
                        extracted_values.append(extracted_value)
                        self.logger.info(
                            f"🔍 Valor extraído del PDF después de '{col_name}': '{extracted_value}'"
                        )

                        is_in_search_keys = extracted_value in self.search_to_rename_map
                        is_in_rename_values = (
                            extracted_value in self.search_to_rename_map.values()
                        )

                        self.logger.info(
                            f"🔍 DEBUG VERIFICACIÓN: "
                            f"¿Está '{extracted_value}' en claves de búsqueda Excel? {is_in_search_keys} "
                            f"¿Está en valores de renombrado Excel? {is_in_rename_values}"
                        )

                        if is_in_search_keys:
                            corresponding_rename = self.search_to_rename_map[
                                extracted_value
                            ]
                            self.logger.info(
                                f"🔍 Si está en claves Excel, correspondería a: '{corresponding_rename}'"
                            )

                        if is_in_rename_values:
                            for (
                                search_key,
                                rename_val,
                            ) in self.search_to_rename_map.items():
                                if rename_val == extracted_value:
                                    self.logger.info(
                                        f"🔍 Si está en valores Excel, viene de la clave: '{search_key}'"
                                    )
                                    break

                        for (
                            excel_key,
                            rename_value,
                        ) in self.search_to_rename_map.items():
                            excel_key_str = str(excel_key).strip()
                            if excel_key_str.lower() == extracted_value.lower():
                                self.logger.info(
                                    f"✅ COINCIDENCIA POR HEADER ENCONTRADA: '{extracted_value}' -> '{rename_value}'"
                                )
                                return rename_value

                        self.logger.debug(
                            "No se encontró coincidencia exacta, buscando coincidencia parcial..."
                        )
                        for (
                            excel_key,
                            rename_value,
                        ) in self.search_to_rename_map.items():
                            excel_key_str = str(excel_key).strip()
                            if (
                                extracted_value.lower() in excel_key_str.lower()
                                or excel_key_str.lower() in extracted_value.lower()
                            ):
                                self.logger.info(
                                    f"✅ COINCIDENCIA PARCIAL POR HEADER ENCONTRADA: '{extracted_value}' ≈ '{excel_key}' -> '{rename_value}'"
                                )
                                return rename_value

        self.logger.debug("=== MÉTODO 3: Búsqueda flexible con palabras clave ===")
        for excel_key, rename_value in list(self.search_to_rename_map.items())[:5]:
            excel_key_normalized = self._normalize_text_for_search(str(excel_key))
            if not excel_key_normalized:
                continue

            key_words = excel_key_normalized.split()
            if len(key_words) >= 2:
                for word in key_words:
                    if len(word) >= 4 and word in normalized_pdf_text:
                        self.logger.info(
                            f"✅ COINCIDENCIA FLEXIBLE ENCONTRADA: palabra '{word}' de '{excel_key}' -> '{rename_value}'"
                        )
                        return rename_value

        if extracted_values:
            fallback_value = extracted_values[0]
            sanitized_name = self._sanitize_filename(fallback_value)
            self.logger.info(
                f"🔄 USANDO VALOR EXTRAÍDO COMO FALLBACK: '{fallback_value}' -> '{sanitized_name}'"
            )
            return sanitized_name

        self.logger.warning("❌ No se encontró ninguna coincidencia en el texto")
        return None

    def _save_debug_text(self, text: str, page_num: int, support_num: int = 1):
        """Guarda el texto extraído para debugging manual"""
        debug_dir = self.output_dir / "debug_texts"
        debug_dir.mkdir(exist_ok=True)

        debug_file = debug_dir / f"page_{page_num}_support_{support_num}_debug.txt"

        try:
            with open(debug_file, "w", encoding="utf-8") as f:
                f.write("=== TEXTO ORIGINAL ===\n")
                f.write(text)
                f.write("\n\n=== TEXTO NORMALIZADO ===\n")
                f.write(self._normalize_text_for_search(text))
                f.write("\n\n=== VALORES DEL EXCEL PARA COMPARAR ===\n")
                for i, (search_key, rename_val) in enumerate(
                    list(self.search_to_rename_map.items())[:10], 1
                ):
                    f.write(
                        f"{i}. Buscar: '{search_key}' -> Renombrar: '{rename_val}'\n"
                    )

            self.logger.debug(f"Texto de debug guardado en: {debug_file}")
        except Exception as e:
            self.logger.warning(f"No se pudo guardar texto de debug: {e}")

    def validate_input(self):
        if not self.input_pdf_path.exists():
            raise FileNotFoundError(f"El archivo PDF no existe: {self.input_pdf_path}")

    def create_output_directory(self):
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def extract_metadata(self, pdf_document):
        """Extrae metadatos básicos del PDF"""
        metadata = pdf_document.metadata
        return {
            "title": metadata.get("title", "Sin título"),
            "author": metadata.get("author", "Desconocido"),
            "subject": metadata.get("subject", "Sin asunto"),
            "creator": metadata.get("creator", "Desconocido"),
            "producer": metadata.get("producer", "Desconocido"),
            "creation_date": metadata.get("creationDate", "Desconocida"),
            "modification_date": metadata.get("modDate", "Desconocida"),
            "total_pages": pdf_document.page_count,
        }

    def detect_payment_supports_advanced(self, page, page_num=0, total_pages=0):
        """
        Detección avanzada de soportes de pago usando análisis de texto y elementos visuales
        """
        text = page.get_text()

        payment_indicators = [
            "comprobante",
            "soporte",
            "pago",
            "transferencia",
            "payment",
            "receipt",
            "voucher",
            "transaction",
            "reference",
            "número de referencia",
            "transaction reference number",
            "numero de transaccion",
            "codigo de transaccion",
            "ref:",
            "reference:",
            "nro:",
            "numero:",
            "fecha:",
            "date:",
            "valor:",
            "amount:",
            "monto:",
            "beneficiario:",
            "destinatario:",
            "cuenta:",
            "account:",
            "banco:",
            "bank:",
        ]

        support_regions = []
        page_height = page.rect.height

        support_height = page_height / 3

        y_positions = [0, support_height, 2 * support_height]

        for i, y_start in enumerate(y_positions):
            y_end = min(y_start + support_height, page_height)

            region_rect = fitz.Rect(0, y_start, page.rect.width, y_end)
            region_text = page.get_textbox(region_rect).lower()

            score = 0
            for indicator in payment_indicators:
                if indicator in region_text:
                    score += 1

            if score >= 2:
                support_regions.append(
                    {
                        "rect": region_rect,
                        "confidence": min(score / len(payment_indicators), 1.0),
                        "region_text": region_text[:200] + "..."
                        if len(region_text) > 200
                        else region_text,
                    }
                )

        if not support_regions:
            support_regions.append(
                {
                    "rect": page.rect,
                    "confidence": 0.1,
                    "region_text": text[:200] + "..." if len(text) > 200 else text,
                }
            )

        return support_regions

    def detect_payment_supports(self, page, page_num=0, total_pages=0):
        """Detecta regiones que parecen ser soportes de pago en una página"""
        try:
            regions = self.detect_payment_supports_advanced(page, page_num, total_pages)

            self.logger.debug(
                f"Página {page_num + 1}: Detectadas {len(regions)} regiones de soportes"
            )
            for i, region in enumerate(regions):
                self.logger.debug(
                    f"  Región {i + 1}: Confianza {region['confidence']:.2f}"
                )

            return regions

        except Exception as e:
            self.logger.error(
                f"Error detectando soportes en página {page_num + 1}: {e}"
            )

            return [
                {
                    "rect": page.rect,
                    "confidence": 0.1,
                    "region_text": f"Error en detección: {str(e)[:100]}",
                }
            ]

    def separate_supports_from_page(self, pdf_document, page_num):
        """Separa los soportes individuales de una página del PDF"""
        page = pdf_document[page_num]

        support_regions = self.detect_payment_supports(
            page, page_num, pdf_document.page_count
        )

        created_files = []

        if len(support_regions) <= 1:
            page_text = page.get_text()

            if not page_text.strip():
                self.logger.warning(f"Página {page_num + 1}: Página vacía, omitiendo")
                return created_files

            self.logger.info(f"Página {page_num + 1}: Procesando como soporte único")

            self._save_debug_text(page_text, page_num + 1, 1)

            rename_value = self._find_rename_in_text(page_text)

            if rename_value:
                output_filename = f"{rename_value}.pdf"
                self.logger.info(
                    f"Página {page_num + 1}: Renombrando a '{output_filename}'"
                )
            else:
                output_filename = f"soporte_pagina_{page_num + 1}.pdf"
                self.logger.warning(
                    f"Página {page_num + 1}: No se encontró valor para renombrar, usando '{output_filename}'"
                )

            output_path = self._unique_path(self.output_dir, output_filename)

            new_doc = fitz.open()
            new_doc.insert_pdf(pdf_document, from_page=page_num, to_page=page_num)
            new_doc.save(str(output_path))
            new_doc.close()

            created_files.append(
                {
                    "file": str(output_path),
                    "page": page_num + 1,
                    "support": 1,
                    "rename_value": rename_value,
                    "confidence": support_regions[0].get("confidence", 0.1)
                    if support_regions
                    else 0.1,
                }
            )

            self.logger.info(f"✅ Creado: {output_path.name}")

        else:
            self.logger.info(
                f"Página {page_num + 1}: Detectados {len(support_regions)} soportes"
            )

            for support_idx, region in enumerate(support_regions, 1):
                try:
                    clip_rect = region["rect"]

                    new_doc = fitz.open()
                    new_page = new_doc.new_page(
                        width=clip_rect.width, height=clip_rect.height
                    )

                    source_area = fitz.Rect(0, 0, clip_rect.width, clip_rect.height)
                    new_page.show_pdf_page(
                        source_area, pdf_document, page_num, clip=clip_rect
                    )

                    clip_text = new_page.get_text()

                    if clip_text.strip():
                        self._save_debug_text(clip_text, page_num + 1, support_idx)

                        rename_value = self._find_rename_in_text(clip_text)

                        if rename_value:
                            output_filename = f"{rename_value}.pdf"
                            self.logger.info(
                                f"Soporte {support_idx}: Renombrando a '{output_filename}'"
                            )
                        else:
                            output_filename = (
                                f"soporte_pagina_{page_num + 1}_parte_{support_idx}.pdf"
                            )
                            self.logger.warning(
                                f"Soporte {support_idx}: No se encontró valor para renombrar, usando '{output_filename}'"
                            )

                        output_path = self._unique_path(
                            self.output_dir, output_filename
                        )

                        new_doc.save(str(output_path))

                        created_files.append(
                            {
                                "file": str(output_path),
                                "page": page_num + 1,
                                "support": support_idx,
                                "rename_value": rename_value,
                                "confidence": region.get("confidence", 0.5),
                            }
                        )

                        self.logger.info(f"✅ Creado: {output_path.name}")
                    else:
                        self.logger.warning(
                            f"Soporte {support_idx}: Región sin texto, omitiendo"
                        )

                    new_doc.close()

                except Exception as e:
                    self.logger.error(f"Error procesando soporte {support_idx}: {e}")
                    if "new_doc" in locals():
                        new_doc.close()

        return created_files

    def separate_pages(self):
        """Separa el PDF en soportes individuales"""
        self.validate_input()
        self.create_output_directory()

        pdf_document = fitz.open(str(self.input_pdf_path))
        total_pages = pdf_document.page_count

        self.logger.info(f"📄 Procesando PDF: {self.input_pdf_path.name}")
        self.logger.info(f"📊 Total de páginas: {total_pages}")
        self.logger.info(f"📁 Directorio de salida: {self.output_dir}")

        all_created_files = []

        for page_num in range(total_pages):
            self.logger.info(f"🔄 Procesando página {page_num + 1}/{total_pages}")

            try:
                created_files = self.separate_supports_from_page(pdf_document, page_num)
                all_created_files.extend(created_files)

                if created_files:
                    self.logger.info(
                        f"✅ Página {page_num + 1}: {len(created_files)} soporte(s) creado(s)"
                    )
                else:
                    self.logger.warning(
                        f"⚠️ Página {page_num + 1}: No se crearon soportes"
                    )

            except Exception as e:
                self.logger.error(f"❌ Error en página {page_num + 1}: {e}")

        pdf_document.close()

        if all_created_files:
            self.logger.info(
                f"🎉 Procesamiento completado: {len(all_created_files)} archivos creados"
            )

            metadata = self.extract_metadata(fitz.open(str(self.input_pdf_path)))
            self.create_summary_report(metadata, all_created_files)
        else:
            self.logger.warning("⚠️ No se crearon archivos de salida")

        return all_created_files

    def create_summary_report(self, metadata, created_files):
        """Crea un informe resumen del procesamiento"""
        try:
            report_path = self.output_dir / "resumen_procesamiento.txt"

            with open(report_path, "w", encoding="utf-8") as f:
                f.write("=== RESUMEN DE PROCESAMIENTO ===\n\n")
                f.write(f"Archivo procesado: {self.input_pdf_path.name}\n")
                f.write(
                    f"Fecha de procesamiento: {self.logger.handlers[0].formatter.formatTime if self.logger.handlers else 'N/A'}\n"
                )
                f.write(f"Total de páginas: {metadata['total_pages']}\n")
                f.write(f"Archivos creados: {len(created_files)}\n\n")

                f.write("=== ARCHIVOS CREADOS ===\n")
                for i, file_info in enumerate(created_files, 1):
                    f.write(f"{i}. {Path(file_info['file']).name}\n")
                    f.write(f"   - Página: {file_info['page']}\n")
                    f.write(f"   - Soporte: {file_info['support']}\n")
                    f.write(
                        f"   - Valor de renombrado: {file_info.get('rename_value', 'N/A')}\n"
                    )
                    f.write(f"   - Confianza: {file_info.get('confidence', 0):.2f}\n\n")

            self.logger.info(f"📋 Resumen guardado en: {report_path}")

        except Exception as e:
            self.logger.error(f"Error creando resumen: {e}")

    def extract_payment_info(self, page, region=None):
        """Extrae información específica de pago de una región de página"""
        if region:
            text = page.get_textbox(region)
        else:
            text = page.get_text()

        payment_info = {
            "reference_number": None,
            "amount": None,
            "date": None,
            "beneficiary": None,
            "account": None,
            "bank": None,
        }

        patterns = {
            "reference_number": [
                r"(?:reference|ref|número|numero|nro)[\s:]*([A-Z0-9]+)",
                r"transaction\s+reference\s+number[\s:]*([A-Z0-9]+)",
                r"codigo[\s:]*([A-Z0-9]+)",
            ],
            "amount": [
                r"(?:valor|amount|monto)[\s:]*([0-9,\.]+)",
                r"\$[\s]*([0-9,\.]+)",
                r"([0-9,\.]+)[\s]*(?:USD|EUR|COP|BRL)",
            ],
            "date": [
                r"(?:fecha|date)[\s:]*([0-9]{1,2}[\/\-][0-9]{1,2}[\/\-][0-9]{2,4})",
                r"([0-9]{1,2}[\/\-][0-9]{1,2}[\/\-][0-9]{2,4})",
            ],
            "beneficiary": [
                r"(?:beneficiario|destinatario|to)[\s:]*([A-Za-z\s]+)",
            ],
            "account": [
                r"(?:cuenta|account)[\s:]*([0-9\-]+)",
            ],
            "bank": [
                r"(?:banco|bank)[\s:]*([A-Za-z\s]+)",
            ],
        }

        for field, field_patterns in patterns.items():
            for pattern in field_patterns:
                matches = re.findall(pattern, text, re.IGNORECASE)
                if matches:
                    payment_info[field] = matches[0].strip()
                    break

        return payment_info

    def create_detailed_summary_report(self, metadata, created_files):
        """Crea un informe detallado con información extraída de cada soporte"""
        try:
            report_data = []

            for file_info in created_files:
                file_path = Path(file_info["file"])

                if file_path.exists():
                    pdf_doc = fitz.open(str(file_path))
                    if pdf_doc.page_count > 0:
                        page = pdf_doc[0]
                        payment_info = self.extract_payment_info(page)

                        report_data.append(
                            {
                                "archivo": file_path.name,
                                "pagina_original": file_info["page"],
                                "soporte_numero": file_info["support"],
                                "valor_renombrado": file_info.get(
                                    "rename_value", "N/A"
                                ),
                                "confianza": file_info.get("confidence", 0),
                                "numero_referencia": payment_info.get(
                                    "reference_number", ""
                                ),
                                "monto": payment_info.get("amount", ""),
                                "fecha": payment_info.get("date", ""),
                                "beneficiario": payment_info.get("beneficiary", ""),
                                "cuenta": payment_info.get("account", ""),
                                "banco": payment_info.get("bank", ""),
                            }
                        )

                    pdf_doc.close()

            if self.export_format == "xlsx" and Workbook is not None:
                self._save_xlsx_report(report_data, metadata)
            else:
                self._save_csv_report(report_data, metadata)

        except Exception as e:
            self.logger.error(f"Error creando informe detallado: {e}")

    def _save_csv_report(self, report_data, metadata):
        """Guarda el informe en formato CSV"""
        report_path = self.output_dir / "informe_detallado.csv"

        import csv

        fieldnames = [
            "archivo",
            "pagina_original",
            "soporte_numero",
            "valor_renombrado",
            "confianza",
            "numero_referencia",
            "monto",
            "fecha",
            "beneficiario",
            "cuenta",
            "banco",
        ]

        with open(report_path, "w", newline="", encoding="utf-8") as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(report_data)

        self.logger.info(f"📊 Informe CSV guardado en: {report_path}")

    def _save_xlsx_report(self, report_data, metadata):
        """Guarda el informe en formato Excel"""
        report_path = self.output_dir / "informe_detallado.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "Soportes Procesados"

        headers = [
            "Archivo",
            "Página Original",
            "Soporte #",
            "Valor Renombrado",
            "Confianza",
            "Número Referencia",
            "Monto",
            "Fecha",
            "Beneficiario",
            "Cuenta",
            "Banco",
        ]

        ws.append(headers)

        for row_data in report_data:
            row = [
                row_data["archivo"],
                row_data["pagina_original"],
                row_data["soporte_numero"],
                row_data["valor_renombrado"],
                row_data["confianza"],
                row_data["numero_referencia"],
                row_data["monto"],
                row_data["fecha"],
                row_data["beneficiario"],
                row_data["cuenta"],
                row_data["banco"],
            ]
            ws.append(row)

        wb.save(str(report_path))
        self.logger.info(f"📊 Informe Excel guardado en: {report_path}")

    def extract_text_from_pages(self):
        """Extrae y guarda el texto de cada página del PDF original"""
        try:
            text_dir = self.output_dir / "textos_extraidos"
            text_dir.mkdir(exist_ok=True)

            pdf_document = fitz.open(str(self.input_pdf_path))

            for page_num in range(pdf_document.page_count):
                page = pdf_document[page_num]
                text = page.get_text()

                if text.strip():
                    text_file = text_dir / f"pagina_{page_num + 1}_texto.txt"

                    with open(text_file, "w", encoding="utf-8") as f:
                        f.write(f"=== TEXTO DE PÁGINA {page_num + 1} ===\n\n")
                        f.write(text)

                    self.logger.debug(
                        f"Texto extraído de página {page_num + 1}: {text_file}"
                    )

            pdf_document.close()
            self.logger.info(f"📝 Textos extraídos guardados en: {text_dir}")

        except Exception as e:
            self.logger.error(f"Error extrayendo textos: {e}")
